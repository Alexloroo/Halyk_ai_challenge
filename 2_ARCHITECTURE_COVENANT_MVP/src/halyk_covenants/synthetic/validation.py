import hashlib
import json
from pathlib import Path

import fitz
from openpyxl import load_workbook

from halyk_covenants.domain import CovenantSpec
from halyk_covenants.synthetic.models import (
    BenchmarkCase,
    DatasetManifest,
    ValidationReport,
)

REQUIRED_SHEETS = ["transactions", "borrowers", "data_dictionary", "known_anomalies"]


class DatasetValidationError(ValueError):
    pass


def validate_dataset(root: Path) -> ValidationReport:
    checks: list[str] = []
    errors: list[str] = []
    manifest_path = root / "manifest.json"
    if not manifest_path.is_file():
        return ValidationReport(valid=False, checks=[], errors=["manifest.json is missing"])

    try:
        manifest = DatasetManifest.model_validate_json(manifest_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return ValidationReport(valid=False, checks=[], errors=[f"manifest is invalid: {exc}"])

    root_resolved = root.resolve()
    for artifact in manifest.artifacts:
        artifact_path = (root / artifact.path).resolve()
        if not artifact_path.is_relative_to(root_resolved):
            errors.append(f"artifact path escapes dataset root: {artifact.path}")
            continue
        if not artifact_path.is_file():
            errors.append(f"artifact is missing: {artifact.path}")
            continue
        digest = hashlib.sha256(artifact_path.read_bytes()).hexdigest()
        if digest != artifact.sha256:
            errors.append(f"artifact hash mismatch: {artifact.path}")
        if artifact_path.stat().st_size != artifact.size_bytes:
            errors.append(f"artifact size mismatch: {artifact.path}")
    checks.append(f"verified {len(manifest.artifacts)} artifact hashes")

    document_names = set(manifest.document_defects)
    for document_name in document_names:
        pdf_path = root / "documents" / document_name
        if not pdf_path.is_file():
            errors.append(f"document is missing: {document_name}")
            continue
        try:
            with fitz.open(pdf_path) as document:
                text = "\n".join(page.get_text() for page in document)
                if document.page_count < 1:
                    errors.append(f"document has no pages: {document_name}")
                normalized_text = text.casefold()
                if "synthetic" not in normalized_text and "синтетическ" not in normalized_text:
                    errors.append(f"document lacks synthetic marker: {document_name}")
        except Exception as exc:
            errors.append(f"document cannot be opened: {document_name}: {exc}")
    checks.append(f"opened {len(document_names)} PDFs and checked native text")

    workbook_path = root / "transactions" / "synthetic_transactions.xlsx"
    borrower_ids: set[str] = set()
    if not workbook_path.is_file():
        errors.append("transaction workbook is missing")
    else:
        try:
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
            if workbook.sheetnames != REQUIRED_SHEETS:
                errors.append(f"unexpected workbook sheets: {workbook.sheetnames}")
            else:
                borrower_ids = {
                    str(row[0].value)
                    for row in workbook["borrowers"].iter_rows(min_row=2)
                    if row[0].value is not None
                }
            workbook.close()
        except Exception as exc:
            errors.append(f"transaction workbook cannot be opened: {exc}")
    checks.append("checked workbook topology and borrower registry")

    covenant_ids: set[str] = set()
    for covenant_path in sorted((root / "covenants").glob("*.json")):
        try:
            covenant = CovenantSpec.model_validate_json(covenant_path.read_text(encoding="utf-8"))
            covenant_ids.add(covenant.covenant_id)
        except Exception as exc:
            errors.append(f"invalid covenant {covenant_path.name}: {exc}")
    checks.append(f"validated {len(covenant_ids)} golden covenants")

    cases_path = root / "benchmark" / "cases.json"
    cases: list[BenchmarkCase] = []
    try:
        case_payloads = json.loads(cases_path.read_text(encoding="utf-8"))
        cases = [BenchmarkCase.model_validate(payload) for payload in case_payloads]
    except Exception as exc:
        errors.append(f"benchmark cases are invalid: {exc}")
    for case in cases:
        if case.covenant_id not in covenant_ids:
            errors.append(f"case {case.case_id} references missing covenant {case.covenant_id}")
        if case.borrower_id not in borrower_ids:
            errors.append(f"case {case.case_id} references missing borrower {case.borrower_id}")
        if case.document_file not in document_names:
            errors.append(f"case {case.case_id} references missing document {case.document_file}")
    checks.append(f"cross-referenced {len(cases)} benchmark cases")

    qa_path = root / "benchmark" / "qa_pairs.jsonl"
    try:
        qa_records = [json.loads(line) for line in qa_path.read_text(encoding="utf-8").splitlines()]
        if [record["case_id"] for record in qa_records] != [case.case_id for case in cases]:
            errors.append("Q&A case order or IDs do not match cases.json")
        for case, record in zip(cases, qa_records, strict=True):
            if record.get("answer") != case.expected.model_dump(mode="json"):
                errors.append(f"Q&A answer does not match case expectation: {case.case_id}")
    except Exception as exc:
        errors.append(f"Q&A JSONL is invalid: {exc}")
    checks.append("cross-referenced machine-readable Q&A")

    return ValidationReport(valid=not errors, checks=checks, errors=errors)


def require_valid_dataset(root: Path) -> ValidationReport:
    report = validate_dataset(root)
    if not report.valid:
        raise DatasetValidationError("; ".join(report.errors))
    return report
