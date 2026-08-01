import os
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from halyk_covenants.synthetic.models import SyntheticDatasetDefinition

TRANSACTION_HEADERS = [
    "transaction_id",
    "borrower_id",
    "account_id",
    "transaction_date",
    "amount",
    "currency",
    "direction",
    "counterparty_id",
    "counterparty_name",
    "purpose",
    "source_row_id",
]

_FIXED_MODIFIED_ELEMENT = (
    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2026-08-02T00:00:00Z</dcterms:modified>'
)


def render_workbook(definition: SyntheticDatasetDefinition, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    fixed_timestamp = datetime(2026, 8, 2, 0, 0, 0)
    workbook.properties.created = fixed_timestamp
    workbook.properties.modified = fixed_timestamp
    transactions = workbook.active
    transactions.title = "transactions"
    transactions.append(TRANSACTION_HEADERS)
    for transaction in definition.transactions:
        transactions.append(
            [
                transaction.transaction_id,
                transaction.borrower_id,
                transaction.account_id,
                transaction.transaction_date,
                _excel_number(transaction.amount),
                transaction.currency,
                transaction.direction,
                transaction.counterparty_id,
                transaction.counterparty_name,
                transaction.purpose,
                transaction.source_row_id,
            ]
        )
    _style_table(transactions, text_columns={1, 2, 3, 8, 11})
    for cell in transactions["E"][1:]:
        cell.number_format = "#,##0.000000"
    for cell in transactions["D"][1:]:
        cell.number_format = "yyyy-mm-dd"

    borrowers = workbook.create_sheet("borrowers")
    borrowers.append(["borrower_id", "canonical_name", "aliases", "synthetic_identifier"])
    for borrower in definition.borrowers:
        borrowers.append(
            [
                borrower.borrower_id,
                borrower.canonical_name,
                "; ".join(borrower.aliases),
                borrower.identifiers["synthetic"],
            ]
        )
    _style_table(borrowers, text_columns={1, 4})

    dictionary = workbook.create_sheet("data_dictionary")
    dictionary.append(["column", "type", "required", "description"])
    descriptions = {
        "transaction_id": ("string", "yes", "Synthetic transaction identifier."),
        "borrower_id": ("string", "yes", "Foreign key to borrowers.borrower_id."),
        "account_id": ("string", "no", "Optional synthetic account identifier."),
        "transaction_date": ("date", "yes", "Operation date in Asia/Almaty calendar terms."),
        "amount": ("decimal(38,6)", "yes", "Transaction amount without FX conversion."),
        "currency": ("string", "yes", "ISO-like currency code used by covenant filters."),
        "direction": ("string", "yes", "incoming or outgoing."),
        "counterparty_id": ("string", "no", "Optional synthetic counterparty ID."),
        "counterparty_name": ("string", "no", "Optional synthetic counterparty name."),
        "purpose": ("string", "no", "Optional payment purpose."),
        "source_row_id": ("string", "yes", "Synthetic source-row provenance."),
    }
    for column in TRANSACTION_HEADERS:
        dictionary.append([column, *descriptions[column]])
    _style_table(dictionary, text_columns={1})

    anomalies = workbook.create_sheet("known_anomalies")
    anomaly_headers = ["anomaly_id", "location", "description", "expected_behavior"]
    anomalies.append(anomaly_headers)
    for anomaly in definition.known_anomalies:
        anomalies.append([anomaly[header] for header in anomaly_headers])
    _style_table(anomalies, text_columns={1})

    workbook.save(path)
    _normalize_xlsx_archive(path)
    return path


def _excel_number(value: Decimal) -> int | float:
    integral = value.to_integral_value()
    return int(integral) if value == integral else float(value)


def _style_table(sheet: Worksheet, *, text_columns: set[int]) -> None:
    header_fill = PatternFill("solid", fgColor="006B57")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if index in text_columns:
                cell.number_format = "@"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)


def _normalize_xlsx_archive(path: Path) -> None:
    temporary = path.with_suffix(".normalized.xlsx")
    with (
        ZipFile(path, "r") as source,
        ZipFile(temporary, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            normalized = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            normalized.compress_type = ZIP_DEFLATED
            normalized.external_attr = original.external_attr
            normalized.create_system = original.create_system
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb'<dcterms:modified xsi:type="dcterms:W3CDTF">[^<]+</dcterms:modified>',
                    _FIXED_MODIFIED_ELEMENT,
                    payload,
                )
            target.writestr(normalized, payload)
    os.replace(temporary, path)
