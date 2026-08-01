from pathlib import Path

import fitz
from openpyxl import load_workbook

from halyk_covenants.synthetic import build_synthetic_definition
from halyk_covenants.synthetic.pdf import render_pdfs
from halyk_covenants.synthetic.workbook import render_workbook


def test_workbook_preserves_sheet_contract_text_ids_numeric_amounts_and_anomalies(
    tmp_path: Path,
) -> None:
    output = tmp_path / "synthetic_transactions.xlsx"

    render_workbook(build_synthetic_definition(), output)

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "transactions",
        "borrowers",
        "data_dictionary",
        "known_anomalies",
    ]
    transactions = workbook["transactions"]
    headers = [cell.value for cell in transactions[1]]
    assert headers == [
        "transaction_id",
        "borrower_id",
        "account_id",
        "transaction_date",
        "amount",
        "currency",
        "direction",
        "counterparty_id",
        "counterparty_name",
        "purpose",
        "source_row_id",
    ]
    assert all(isinstance(transactions.cell(row=row, column=1).value, str) for row in range(2, 16))
    assert all(isinstance(transactions.cell(row=row, column=2).value, str) for row in range(2, 16))
    assert all(
        isinstance(transactions.cell(row=row, column=5).value, (int, float)) for row in range(2, 16)
    )
    assert transactions.cell(row=5, column=1).value == "000001"
    assert transactions.cell(row=5, column=1).number_format == "@"
    assert transactions.freeze_panes == "A2"
    assert transactions.auto_filter.ref == "A1:K15"

    duplicate_rows = [
        tuple(cell.value for cell in transactions[row])
        for row in range(2, transactions.max_row + 1)
        if transactions.cell(row=row, column=1).value == "000003"
    ]
    assert len(duplicate_rows) == 2
    assert duplicate_rows[0] == duplicate_rows[1]
    assert workbook["known_anomalies"].max_row == 5


def test_pdf_renderer_creates_native_cyrillic_text_and_real_table_grids(tmp_path: Path) -> None:
    output_dir = tmp_path / "documents"

    paths = render_pdfs(build_synthetic_definition(), output_dir)

    assert [path.name for path in paths] == [
        "alpha_trade_contract.pdf",
        "borrower_limits_appendix.pdf",
    ]
    alpha = fitz.open(paths[0])
    appendix = fitz.open(paths[1])
    try:
        assert alpha.page_count >= 1
        assert appendix.page_count == 2
        alpha_text = "\n".join(page.get_text() for page in alpha)
        appendix_text = "\n".join(page.get_text() for page in appendix)
        assert "Альфа Трейд" in alpha_text
        assert "обьем" in alpha_text
        assert "COV-ALPHA-SUM" in alpha_text
        assert "Beta Logistics" in appendix_text
        assert "000777" in appendix_text
        assert "млн KZT" in appendix_text
        assert "Исключение" in appendix_text
        assert sum(len(page.get_drawings()) for page in alpha) >= 10
        assert sum(len(page.get_drawings()) for page in appendix) >= 15
    finally:
        alpha.close()
        appendix.close()
