import json
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from halyk_covenants.synthetic import build_synthetic_definition
from halyk_covenants.synthetic.generator import generate_synthetic_dataset
from halyk_covenants.synthetic.models import DatasetManifest
from halyk_covenants.synthetic.validation import DatasetValidationError, validate_dataset

EXPECTED_COVENANTS = {
    "COV-ALPHA-SUM.json",
    "COV-ALPHA-MAX.json",
    "COV-ALPHA-COUNT.json",
    "COV-ALPHA-MIN.json",
    "COV-BETA-AVG.json",
    "COV-BETA-SUM.json",
    "COV-BETA-MAX.json",
    "COV-GAMMA-SUM.json",
}


def test_generator_creates_complete_validated_artifact_topology(tmp_path: Path) -> None:
    output = tmp_path / "synthetic"

    manifest = generate_synthetic_dataset(output)

    assert isinstance(manifest, DatasetManifest)
    assert {path.name for path in (output / "documents").iterdir()} == {
        "alpha_trade_contract.pdf",
        "borrower_limits_appendix.pdf",
    }
    assert {path.name for path in (output / "covenants").iterdir()} == EXPECTED_COVENANTS
    assert (output / "transactions" / "synthetic_transactions.xlsx").is_file()
    assert {path.name for path in (output / "benchmark").iterdir()} == {
        "cases.json",
        "qa_pairs.jsonl",
        "qa_pairs.md",
    }
    assert (output / "manifest.json").is_file()
    assert len(manifest.artifacts) == 14
    assert validate_dataset(output).valid is True


def test_q_and_a_and_cases_are_cross_referenced_to_golden_artifacts(tmp_path: Path) -> None:
    output = tmp_path / "synthetic"
    generate_synthetic_dataset(output)

    cases = json.loads((output / "benchmark" / "cases.json").read_text(encoding="utf-8"))
    qa_records = [
        json.loads(line)
        for line in (output / "benchmark" / "qa_pairs.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
    ]

    assert len(cases) == len(qa_records) == 10
    assert [case["case_id"] for case in cases] == [record["case_id"] for record in qa_records]
    assert all((output / "covenants" / f"{case['covenant_id']}.json").is_file() for case in cases)
    assert all(
        record["answer"] == case["expected"] for case, record in zip(cases, qa_records, strict=True)
    )
    assert "ALPHA-COUNT-TRIGGER" in (output / "benchmark" / "qa_pairs.md").read_text(
        encoding="utf-8"
    )

    workbook = load_workbook(output / "transactions" / "synthetic_transactions.xlsx")
    borrower_ids = {str(row[0].value) for row in workbook["borrowers"].iter_rows(min_row=2)}
    assert {case["borrower_id"] for case in cases} <= borrower_ids


def test_manifest_hashes_are_deterministic_across_independent_generations(tmp_path: Path) -> None:
    first = generate_synthetic_dataset(tmp_path / "first")
    second = generate_synthetic_dataset(tmp_path / "second")

    first_hashes = {artifact.path: artifact.sha256 for artifact in first.artifacts}
    second_hashes = {artifact.path: artifact.sha256 for artifact in second.artifacts}
    assert first_hashes == second_hashes


def test_workbook_core_timestamps_are_fixed(tmp_path: Path) -> None:
    output = tmp_path / "synthetic"
    generate_synthetic_dataset(output)

    workbook = output / "transactions" / "synthetic_transactions.xlsx"
    with ZipFile(workbook) as archive:
        core_properties = archive.read("docProps/core.xml").decode("utf-8")

    assert core_properties.count("2026-08-02T00:00:00Z") == 2


def test_invalid_generation_does_not_replace_last_valid_dataset(tmp_path: Path) -> None:
    output = tmp_path / "synthetic"
    generate_synthetic_dataset(output)
    original_manifest = (output / "manifest.json").read_bytes()
    invalid_definition = build_synthetic_definition()
    invalid_definition.cases[0].covenant_id = "COV-DOES-NOT-EXIST"

    with pytest.raises(DatasetValidationError):
        generate_synthetic_dataset(output, definition=invalid_definition)

    assert (output / "manifest.json").read_bytes() == original_manifest
    assert validate_dataset(output).valid is True
